# -*- coding: utf-8 -*-
"""Дашборд: сводка ошибок и выгрузка отчётов по территориям стоп-фактора."""
import io

from openpyxl import Workbook

from services.report_check import STOP_FACTOR_REGIONS, STOP_FACTOR_DISTRICTS

# Виды работ, которые учитываются в дашборде/отчётах (остальные не важны).
DASHBOARD_WORK_TYPES = [
    "Бытовые заявки",
    "План лестница",
    "План квартира",
    "Периодический контроль БП",
    "Допуск ПУ в МКД",
    "Допуск ПУ в ИЖС",
    "План ИЖС",
    "Выявление безучетного потребления БП",
    "Инструментальная проверка",
    "Контроль СП",
]

# Расценки (₽) по видам работ — для плашки «Стоимость» на вкладке «Обзор».
WORK_TYPE_RATES = {
    "Инструментальная проверка": 2967.01,
    "Бытовые заявки": 1000.06,
    "Периодический контроль БП": 657.23,
    "Выявление безучетного потребления БП": 401.71,
    "Допуск ПУ в ИЖС": 819.11,
    "Допуск ПУ в МКД": 256.83,
}


def build_scope(user, dept: str = "") -> tuple[list, dict]:
    """Зона видимости: территории стоп-фактора + 10 видов работ + (отделение) + видимость по роли (все заказчики)."""
    clauses: list[str] = []
    params: dict = {}

    if dept:
        clauses.append("executor_organization = :dept_filter")
        params["dept_filter"] = dept

    wt_names = [f"wt{i}" for i in range(len(DASHBOARD_WORK_TYPES))]
    params.update(zip(wt_names, DASHBOARD_WORK_TYPES))
    wt_in = ", ".join(f":{n}" for n in wt_names)
    clauses.append(f"task_report IN ({wt_in})")

    regions = sorted(STOP_FACTOR_REGIONS)
    districts = sorted(STOP_FACTOR_DISTRICTS)
    r_names = [f"sfr{i}" for i in range(len(regions))]
    d_names = [f"sfd{i}" for i in range(len(districts))]
    params.update(zip(r_names, regions))
    params.update(zip(d_names, districts))
    r_in = ", ".join(f":{n}" for n in r_names)
    d_in = ", ".join(f":{n}" for n in d_names)
    clauses.append(f"(region IN ({r_in}) OR municipal_district IN ({d_in}))")

    role = user.effective_role
    if role in ("оператор", "работник"):
        clauses.append("executor IN (SELECT full_name FROM users WHERE locale = :locale)")
        params["locale"] = user.locale
    elif role == "менеджер":
        clauses.append("executor_organization = :dept")
        params["dept"] = user.dept

    return clauses, params


def pick_pu_type(meter_type_2, meter_type_1, meter_type):
    for v in (meter_type_2, meter_type_1, meter_type):
        if v and str(v).strip():
            return str(v).strip()
    return ""


def generate_errors_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ошибки"
    ws.append(["Номер задания", "Ошибки"])
    for tn, errors in rows:
        ws.append([tn, errors])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_balance_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Балансовая принадлежность"
    ws.append(["Номер задания", "Тип ПУ"])
    for tn, pu_type in rows:
        ws.append([tn, pu_type])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_task_numbers_xlsx(rows) -> bytes:
    """Отчёт из одного столбца «Номер задания» (для «Дата работ» и «Отметка о проверке»)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Задания"
    ws.append(["Номер задания"])
    for tn in rows:
        ws.append([tn])
    ws.column_dimensions["A"].width = 28
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
