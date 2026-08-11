import io as io_module
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy import text

DEPT_PREFIXES = {
    "Приозерское отделение": "Пр", "Лужское отделение": "Лг", "Выборгское отделение": "Вб",
    "Всеволожское отделение": "Вс", "Новоладожское отделение": "Нл", "Кингисеппское отделение": "Кн",
    "Сертоловское отделение": "Ср", "Гатчинское отделение": "Гт", "Тихвинское отделение": "Тх",
    "Курортное отделение": "Кр", "Петродворцовое отделение": "Пт", "Пушкинское отделение": "Пш",
    "Тосненское отделение": "Тс",
}

LOCALE_SUFFIXES = {
    "Всеволожск": "",
    "Выборг": "В",
    "Рощино": "Р",
    "Гатчина": "",
    "Волосово": "В",
    "Кингисепп": "К",
    "Сестрорецк": "",
    "Луга": "",
    "Ладога": "Н",
    "Поле": "Л",
    "Подпорожье": "П",
    "Петергоф": "",
    "Приозерск": "П",
    "Сосново": "С",
    "Пушкин": "",
    "Сертолово": "",
    "Кириши": "К",
    "Тихвин": "Т",
    "Кировск": "К",
    "Тосно": "Т",
}

async def generate_reestr_xlsx_bytes(db_session, task_numbers, reestr_number, reestr_date, task_report, dept, current_user):
    placeholders = ','.join([f"'{tn}'" for tn in task_numbers])
    
    result = await db_session.execute(
        text(f"""SELECT task_number, personal_account, municipal_district, address, done_day
                 FROM main_afl WHERE task_number IN ({placeholders})
                 ORDER BY task_number""")
    )
    data_rows = [dict(row._mapping) for row in result]
    
    # Начальник отделения
    chief_result = await db_session.execute(
        text("SELECT full_name FROM users WHERE dept = :dept AND position = 'Начальник отделения' LIMIT 1"),
        {"dept": dept}
    )
    chief = chief_result.scalar_one_or_none()
    if not chief:
        chief = current_user.full_name
    
    # Форматирование даты
    reestr_date_formatted = ""
    if reestr_date:
        try:
            dt = datetime.strptime(reestr_date, "%Y-%m-%d")
            reestr_date_formatted = dt.strftime("%d.%m.%Y")
        except ValueError:
            reestr_date_formatted = reestr_date
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр"
    
    ws.page_setup.orientation = 'landscape'
    ws.page_margins.left = 1 / 2.54  # 1 см в дюймах
    ws.page_margins.right = 1 / 2.54
    ws.page_margins.top = 2 / 2.54
    ws.page_margins.bottom = 2 / 2.54
    ws.oddFooter.center.text = "Страница &P из &N"
    ws.sheet_properties.pageSetUpPr = None
    
    # Колонки A-G (без дырки)
    col_widths = {'A': 5, 'B': 22, 'C': 18, 'D': 32, 'E': 43, 'F': 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row = 1
    
    # Строка 1: номер и дата
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f'Реестр № {reestr_number} от {reestr_date_formatted}'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    # Строка 2: заголовок
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'передачи первичных документов по результатам выполненных работ'
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    # Строка 3: поручение
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f'по поручению: {task_report}'
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2
    
    # Строка 5: тарифная группа
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'Тарифная группа: население'
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    row += 2
    
    # Заголовки таблицы (A-G)
    headers = ['№ п/п', 'Основание', 'Абонентский номер', 'Административный район', 'Адрес объекта', 'Дата акта']
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    
    for header, col in zip(headers, cols):
        cell = ws[f'{col}{row}']
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    row += 1
    
    # Данные
    for i, d in enumerate(data_rows, 1):
        values = [
            i,
            d.get('task_number', ''),
            d.get('personal_account', ''),
            d.get('municipal_district', ''),
            d.get('address', ''),
            d.get('done_day', ''),
        ]
        for val, col in zip(values, cols):
            cell = ws[f'{col}{row}']
            cell.value = val
            cell.font = Font(size=9)
            cell.border = thin_border
            if col in ('A', 'B', 'C', 'F'):
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 2
    
    # Дата передачи
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'Дата передачи: ________________'
    ws[f'A{row}'].font = Font(size=10)
    row += 2
    
    # Направил / Принял
    ws[f'A{row}'] = 'Направил:'
    ws[f'A{row}'].font = Font(size=9)
    ws[f'E{row}'] = 'Принял:'
    ws[f'E{row}'].font = Font(size=9)
    row += 1
    
    ws[f'A{row}'] = 'Начальник отделения'
    ws[f'A{row}'].font = Font(size=9)
    ws[f'E{row}'] = 'Представитель заказчика'
    ws[f'E{row}'].font = Font(size=9)
    row += 1
    
    ws[f'A{row}'] = chief
    ws[f'A{row}'].font = Font(size=9)
    ws[f'E{row}'].border = bottom_border
    
    output = io_module.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


async def generate_report_xlsx_bytes(db_session, period):
    """Генерирует xlsx-файл отчёта за период"""
    
    columns = [
        "task_number", "task_source", "task_type", "work_type_in_task", "created_at",
        "address", "municipal_district", "house_type", "personal_account",
        "contract_status", "contract_created_at", "debt_amount", "debt_calculation_date",
        "disconnection_reconnection_debt_amount", "debt_calculation_date_1",
        "notification_debt_amount", "debt_calculation_date_2", "due_date",
        "service_object_type", "subscriber_name", "subscriber_type", "subscriber_phone",
        "metering_point", "meter_type", "meter_model", "meter_serial_number",
        "manufacture_year", "last_verification_date", "meter_tariff_rate",
        "integer_capacity", "fractional_capacity", "calibration_interval",
        "rated_current", "rated_voltage", "meter_installation_place", "meter_status",
        "meter_ownership", "active_disconnection", "last_readings_t1_estimated",
        "last_readings_t2_estimated", "last_readings_t3_estimated",
        "last_estimated_readings_date", "last_readings_t1_control",
        "last_readings_t2_control", "last_readings_t3_control",
        "last_control_readings_date", "seals", "has_current_transformer",
        "has_voltage_transformer", "meter_inspection_results", "meter_type_1",
        "meter_model_1", "meter_serial_number_1", "manufacture_year_1",
        "last_verification_date_1", "meter_tariff_rate_1", "integer_capacity_1",
        "fractional_capacity_1", "calibration_interval_1", "rated_current_1",
        "rated_voltage_1", "meter_installation_place_1", "meter_ownership_1",
        "seals_1", "t1", "t2", "t3", "violations", "meter_status_date",
        "meter_malfunction", "unauthorized_interference", "unauthorized_connection",
        "additional_violations", "unsuccessful_inspection_reason", "work_type",
        "work_result", "meter_type_2", "meter_model_2", "meter_serial_number_2",
        "manufacture_year_2", "last_verification_date_2", "meter_tariff_rate_2",
        "integer_capacity_2", "fractional_capacity_2", "calibration_interval_2",
        "rated_current_2", "rated_voltage_2", "meter_installation_place_2",
        "meter_ownership_2", "seals_2", "t1_1", "t2_1", "t3_1", "final_meter_status",
        "acts", "comment", "work_start_date", "work_end_date", "sent_to_billing",
        "billing_sent_at", "task_organization", "third_party_organization",
        "assignee", "executor", "executor_organization", "visit_reason", "verified",
        "status", "work_result_1", "status_changed_at", "task_link",
        "customer", "task_output", "task_report", "task_detail", "region", "grid",
        "reestr_number", "reestr_date"
    ]
    
    headers = [
        "Номер задания", "Источник задания", "Вид задания", "Вид работы в задании",
        "Дата создания", "Адрес", "Муниципальный район", "Тип дома", "Лицевой счет",
        "Статус договора", "Дата создания договора", "Сумма ДЗ", "Дата расчета задолженности",
        "Сумма задолженности за отключение/возобновление ЭЭ", "Дата расчета задолженности.1",
        "Сумма задолженности из уведомления", "Дата расчета задолженности.2", "Срок исполнения",
        "Тип объекта обслуживания", "Наименование абонента", "Тип абонента", "Телефон абонента",
        "Точка учета", "Тип ПУ", "Модель ПУ", "Заводской номер ПУ", "Год выпуска",
        "Дата последней поверки", "Тарифность ПУ", "Разрядность целой части",
        "Разрядность дробной части", "МПИ", "Номинальный ток", "Номинальное напряжение",
        "Место установки ПУ", "Статус ПУ", "Балансовая принадлежность ПУ", "Активное отключение",
        "Последние показания Т1 (расчетные)", "Последние показания Т2 (расчетные)",
        "Последние показания Т3 (расчетные)", "Дата последних расчетных показаний",
        "Последние показания Т1 (контрольные)", "Последние показания Т2 (контрольные)",
        "Последние показания Т3 (контрольные)", "Дата последних контрольных показаний",
        "Пломбы", "Есть трансформатор тока", "Есть трансформатор напряжения",
        "Результаты осмотра/проверки ПУ", "Тип ПУ.1", "Модель ПУ.1", "Заводской номер ПУ.1",
        "Год выпуска.1", "Дата последней поверки.1", "Тарифность ПУ.1",
        "Разрядность целой части.1", "Разрядность дробной части.1", "МПИ.1",
        "Номинальный ток.1", "Номинальное напряжение.1", "Место установки ПУ.1",
        "Балансовая принадлежность ПУ.1", "Пломбы.1", "T1", "T2", "T3", "Нарушения",
        "Дата состояния ПУ", "Неисправность ПУ", "Несанкционированное вмешательство",
        "Несанкционированное подключение", "Доп. нарушения", "Причина нерезультативного осмотра",
        "Вид работы", "Результат работы", "Тип ПУ.2", "Модель ПУ.2", "Заводской номер ПУ.2",
        "Год выпуска.2", "Дата последней поверки.2", "Тарифность ПУ.2",
        "Разрядность целой части.2", "Разрядность дробной части.2", "МПИ.2",
        "Номинальный ток.2", "Номинальное напряжение.2", "Место установки ПУ.2",
        "Балансовая принадлежность ПУ.2", "Пломбы.2", "T1.1", "T2.1", "T3.1",
        "Статус ПУ (итог)", "Акты", "Комментарий", "Дата начала выполнения работы",
        "Дата окончания выполнения работы", "Отправлено в биллинг", "Дата и время отправки в биллинг",
        "Организация задания", "Сторонняя организация", "Ответственный за исполнение",
        "Исполнитель", "Организация исполнителя", "Основание посещения", "Проверено",
        "Статус", "Результат работы.1", "Дата изменения статуса", "Ссылка на задание", "Заказчик",
        "Анализ", "Отчёт", "Расшифровка", "Субъект", "Сетевая", "Номер реестра", "Дата реестра"
    ]
    
    columns_str = ", ".join(columns)
    result = await db_session.execute(
        text(f"SELECT {columns_str} FROM story_afl WHERE report = :period ORDER BY task_number"),
        {"period": period}
    )
    rows = [dict(row._mapping) for row in result]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    
    ws.page_setup.orientation = 'landscape'
    
    # Заголовки
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True, size=8)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Данные
    for r, row_data in enumerate(rows, 2):
        for c, col in enumerate(columns, 1):
            val = row_data.get(col, '')
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(size=8)
            cell.border = thin_border
    
    # Автоширина
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    output = io_module.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()